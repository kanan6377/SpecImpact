from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook

from specimpact.core import analyze_change
from specimpact.models import Artifact, Evidence
from specimpact.reports import export_report_excel
from specimpact.store import LocalStore
from specimpact.tabular_loaders import ingest_excel, inspect_excel_folder
from specimpact.webui.registry import ProjectRegistry
from specimpact.webui.services import execute, project_overview, report_data

ROOT = Path(__file__).parents[1]
SAMPLE = ROOT / "examples" / "japanese_sier_excel"


def test_japanese_sier_excel_mvp_flow(tmp_path: Path) -> None:
    store = LocalStore(tmp_path / ".specimpact")
    records = ingest_excel(store, SAMPLE / "docs", SAMPLE / "aliases.yml")
    health = inspect_excel_folder(SAMPLE / "docs")
    assert len(records) == 6
    assert health["workbooks"] == 6
    assert health["sheets"] == 6

    report = analyze_change(
        store,
        next((SAMPLE / "changes").glob("*.md")),
        no_llm=True,
    )
    grouped = report.grouped()
    must_ids = {item["artifact_id"] for item in grouped["must_review"]}
    assert {
        "screen.enrollment_application",
        "api.enrollment_application",
        "column.requested_credit_limit",
        "validation.credit_limit_upper_bound",
        "external_if.credit_screening",
        "test.credit_limit_boundary",
    } <= must_ids
    assert "screen.item_7bcfde24d4" not in must_ids

    evidence = store.read("evidence", Evidence)
    assert any("[入会申込API!F3]" in item.quote for item in evidence)
    assert any("[CREDIT_APPLICATION!C4]" in item.quote for item in evidence)
    report_path = export_report_excel(store)
    assert report_path.is_file()
    workbook = load_workbook(report_path)
    headers = [cell.value for cell in workbook["ReviewCandidates"][1]]
    assert {"primary_evidence", "evidence_count", "evidence_ids"} <= set(headers)
    api_row = next(
        row for row in workbook["ReviewCandidates"].iter_rows(values_only=True)
        if row[2] == "入会申込API"
    )
    assert api_row[headers.index("evidence_count")] >= 1
    assert str(api_row[headers.index("evidence_ids")]).startswith("ev.")


def test_sier_sheet_detection_uses_headers_when_filename_is_generic(tmp_path: Path) -> None:
    source = SAMPLE / "docs" / "画面設計書.xlsx"
    workbook_path = tmp_path / "別紙2.xlsx"
    shutil.copyfile(source, workbook_path)
    store = LocalStore(tmp_path / ".specimpact")
    ingest_excel(store, workbook_path, SAMPLE / "aliases.yml")
    assert any(
        item.artifact_type == "Screen" and item.display_name == "入会申込画面"
        for item in store.read("artifacts", Artifact)
    )


def test_gui_excel_health_and_change_text_flow(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    shutil.copytree(SAMPLE, workspace)
    project = ProjectRegistry(tmp_path / "registry").add(workspace)
    execute(project, "init", {})
    execute(project, "ingest_excel", {"path": "docs", "aliases": "aliases.yml"})
    overview = project_overview(project)
    assert overview["health_check"]["workbooks"] == 6

    change_text = next((workspace / "changes").glob("*.md")).read_text(encoding="utf-8")
    result = execute(project, "analyze_text", {"body": change_text})
    assert result["result"]["candidates"] >= 6
    report = report_data(project)
    assert any(item["display_name"] == "入会申込画面" for item in report["must_review"])
