from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill

from specimpact.dirty_excel.ingestion import ingest_dirty_excel
from specimpact.dirty_excel.region_detector import detect_regions
from specimpact.dirty_excel.sheet_classifier import classify_sheets
from specimpact.dirty_excel.workbook_reader import read_dirty_workbook
from specimpact.graphrag import FakeLLMClient
from specimpact.impact_management.change_atoms import ChangeAtom, parse_change_atoms
from specimpact.impact_management.decision_store import set_impact_status
from specimpact.impact_management.impact_hypothesis import build_impact_hypotheses
from specimpact.impact_management.impact_retrieval import RetrievedPath
from specimpact.impact_management.review_session import analyze_change_llm_first
from specimpact.llm_graph.entity_resolution import (
    decide_alias_candidate,
    suggest_alias_candidates,
)
from specimpact.llm_graph.extraction import extract_region_with_llm
from specimpact.llm_graph.schemas import AliasCandidate
from specimpact.llm_graph.verifier import classify_impact
from specimpact.models import Artifact, Entity, Evidence, EvidenceSupport, Relation, SourceLocation
from specimpact.store import LocalStore
from specimpact.webui.registry import ProjectRegistry
from specimpact.webui.services import dirty_excel_data, execute, impact_decisions_data

ROOT = Path(__file__).parents[1]
SIER = ROOT / "examples" / "japanese_sier_excel"


class RecordingFakeLLMClient(FakeLLMClient):
    def __init__(self, responses=None, model: str = "fake-model") -> None:
        super().__init__(responses, model)
        self.calls = []

    def structured(self, purpose, payload, schema):
        self.calls.append({"purpose": purpose, "payload": payload, "schema": schema.__name__})
        return super().structured(purpose, payload, schema)


def test_dirty_workbook_normalization_preserves_cells_and_regions(tmp_path: Path) -> None:
    workbook_path = tmp_path / "dirty.xlsx"
    _write_dirty_workbook(workbook_path)
    store = LocalStore(tmp_path / ".specimpact")

    summary = ingest_dirty_excel(store, workbook_path)

    assert summary.workbooks == 1
    assert summary.cells >= 10
    assert summary.regions >= 2
    assert (store.root / "sources" / "original").is_dir()
    normalized = next((store.root / "sources" / "normalized").glob("*.jsonl"))
    assert normalized.is_file()

    _workbook, sheets, cells = read_dirty_workbook(workbook_path)
    assert any(cell.merged_range == "A1:B1" for cell in cells)
    assert any(cell.style.font_bold for cell in cells)
    assert any(cell.style.fill_color for cell in cells)
    assert any(cell.comment == "owner note" for cell in cells)
    assert any(cell.hyperlink == "https://example.com/spec" for cell in cells)
    assert any(cell.is_hidden_row for cell in cells)
    assert any(cell.is_hidden_col for cell in cells)
    assert any(sheet.chart_count for sheet in sheets)
    assert read_dirty_workbook(workbook_path)[0].warnings

    regions = detect_regions(classify_sheets(sheets, cells), cells)
    region_types = {region.region_type for region in regions}
    assert {"revision_history", "screen_item_table", "validation_block"} <= region_types


def test_region_llm_extraction_keeps_only_valid_evidence(tmp_path: Path) -> None:
    workbook_path = tmp_path / "dirty.xlsx"
    _write_dirty_workbook(workbook_path)
    _workbook, sheets, cells = read_dirty_workbook(workbook_path)
    sheets = classify_sheets(sheets, cells)
    region = next(
        region
        for region in detect_regions(sheets, cells)
        if region.region_type == "screen_item_table"
    )
    evidence_id = region.evidence_ids[0]
    valid = {
        "region_id": region.region_id,
        "nodes": [
            {
                "temp_id": "screen",
                "node_type": "Screen",
                "display_name": "入会申込画面",
                "canonical_hint": None,
                "aliases": [],
                "properties": {},
                "evidence_ids": [evidence_id],
                "rationale": "LLM extracted from heading",
            },
            {
                "temp_id": "field",
                "node_type": "ScreenField",
                "display_name": "利用限度額",
                "canonical_hint": None,
                "aliases": [],
                "properties": {},
                "evidence_ids": [evidence_id],
                "rationale": "LLM extracted from row",
            },
        ],
        "edges": [
            {
                "temp_id": "edge",
                "source_temp_id": "screen",
                "relation_type": "DISPLAYS",
                "target_temp_id": "field",
                "evidence_ids": [evidence_id],
                "inference_level": "explicit",
                "rationale": "explicit row",
            }
        ],
        "unresolved_mentions": [],
        "warnings": [],
    }
    invalid = {
        **valid,
        "nodes": [{**valid["nodes"][0], "evidence_ids": ["missing"]}],
        "edges": [],
    }

    client = RecordingFakeLLMClient({"dirty_excel_region_extraction": [valid, invalid]})
    result = extract_region_with_llm(region, cells, client)
    assert len(result.nodes) == 2
    assert len(result.edges) == 1
    payload = client.calls[0]["payload"]
    assert payload["system_prompt"] == payload["instruction"] == payload["instructions"]
    assert payload["prompt"] == payload["system_prompt"]
    assert "screen item definition table" in payload["system_prompt"]
    assert payload["region_type_hint"] == "screen_item_table"
    cleaned = extract_region_with_llm(
        region,
        cells,
        FakeLLMClient({"dirty_excel_region_extraction": invalid}),
    )
    assert cleaned.nodes == []
    assert "invalid evidence_ids missing" in " ".join(cleaned.warnings)


def test_alias_inference_confirm_and_reject_persist(tmp_path: Path) -> None:
    store = LocalStore(tmp_path / ".specimpact")
    store.init()
    store.write(
        "entities",
        [
            Entity(
                entity_id="entity.credit_limit.jp",
                entity_type="BusinessField",
                display_name="利用限度額",
                canonical_name="credit_limit_jp",
            ),
            Entity(
                entity_id="entity.credit_limit.api",
                entity_type="BusinessField",
                display_name="requestedCreditLimit",
                canonical_name="requested_credit_limit",
            ),
            Entity(
                entity_id="entity.credit_limit.db",
                entity_type="BusinessField",
                display_name="REQUESTED_CREDIT_LIMIT",
                canonical_name="requested_credit_limit",
            ),
        ],
    )

    client = RecordingFakeLLMClient(
        {
            "alias_resolution": {
                "judgement": "same",
                "reason": "same credit limit field across screen, API, and DB names",
                "evidence_ids": [],
            }
        }
    )
    assert suggest_alias_candidates(
        store,
        use_llm=True,
        llm_client=client,
    ) == 3
    assert client.calls[0]["payload"]["candidate_signals"]
    candidate = next(
        item
        for item in store.read("alias_candidates", AliasCandidate)
        if item.entity_a_id == "entity.credit_limit.jp"
        and item.entity_b_id == "entity.credit_limit.api"
    )
    assert candidate.judgement == "same"
    assert candidate.entity_a_id == "entity.credit_limit.jp"
    assert candidate.compared_entity_ids
    assert "same credit limit" in candidate.llm_reason
    decide_alias_candidate(store, candidate.candidate_id, "confirmed")
    aliases = (store.root / "aliases.yml").read_text(encoding="utf-8")
    assert "requestedCreditLimit" in aliases
    decide_alias_candidate(store, candidate.candidate_id, "rejected")
    assert (
        next(
            item
            for item in store.read("alias_candidates", AliasCandidate)
            if item.candidate_id == candidate.candidate_id
        ).status
        == "rejected"
    )


def test_alias_llm_judgement_can_reject_similarity(tmp_path: Path) -> None:
    store = LocalStore(tmp_path / ".specimpact")
    store.init()
    store.write(
        "entities",
        [
            Entity(
                entity_id="entity.limit",
                entity_type="BusinessField",
                display_name="requestedCreditLimit",
                canonical_name="credit_limit",
            ),
            Entity(
                entity_id="entity.limit.related",
                entity_type="BusinessField",
                display_name="limitReason",
                canonical_name="credit_limit_reason",
            ),
        ],
    )
    count = suggest_alias_candidates(
        store,
        use_llm=True,
        llm_client=FakeLLMClient(
            {
                "alias_resolution": {
                    "judgement": "different",
                    "reason": "limit amount and reason are different fields",
                    "evidence_ids": [],
                }
            }
        ),
    )
    assert count == 1
    candidate = store.read("alias_candidates", AliasCandidate)[0]
    assert candidate.judgement == "different"
    assert "different fields" in candidate.llm_reason


def test_alias_candidate_recall_uses_name_shape_relations_and_evidence(tmp_path: Path) -> None:
    store = LocalStore(tmp_path / ".specimpact")
    store.init()
    entities = [
        Entity(
            entity_id="entity.postal.camel",
            entity_type="BusinessField",
            display_name="postalCode",
            canonical_name="postal_code",
        ),
        Entity(
            entity_id="entity.postal.snake",
            entity_type="BusinessField",
            display_name="POSTAL_CODE",
            canonical_name="address_postal_code",
        ),
        Entity(
            entity_id="entity.zip.form",
            entity_type="BusinessField",
            display_name="zipFormValue",
            canonical_name="zip_form_value",
        ),
        Entity(
            entity_id="entity.zip.api",
            entity_type="BusinessField",
            display_name="zipCd",
            canonical_name="zip_code",
        ),
    ]
    evidence = [
        Evidence(
            evidence_id="ev.zip.form",
            document_id="doc",
            section_id="sec",
            chunk_id="chunk",
            quote="zipFormValue appears near zipCd mapping",
            evidence_type="test",
            supports=[],
            source_location=SourceLocation(file="spec.xlsx", line_start=10, line_end=10),
        ),
        Evidence(
            evidence_id="ev.zip.api",
            document_id="doc",
            section_id="sec",
            chunk_id="chunk",
            quote="zipCd appears near zipFormValue mapping",
            evidence_type="test",
            supports=[],
            source_location=SourceLocation(file="spec.xlsx", line_start=12, line_end=12),
        ),
    ]
    relations = [
        Relation(
            relation_id="rel.postal.camel",
            relation_type="VALIDATES",
            source_id="validation.address",
            target_id="entity.postal.camel",
            evidence_ids=[],
        ),
        Relation(
            relation_id="rel.postal.snake",
            relation_type="VALIDATES",
            source_id="validation.address",
            target_id="entity.postal.snake",
            evidence_ids=[],
        ),
        Relation(
            relation_id="rel.zip.form",
            relation_type="REQUEST_FIELD",
            source_id="api.address",
            target_id="entity.zip.form",
            evidence_ids=["ev.zip.form"],
        ),
        Relation(
            relation_id="rel.zip.api",
            relation_type="REQUEST_FIELD",
            source_id="api.address",
            target_id="entity.zip.api",
            evidence_ids=["ev.zip.api"],
        ),
    ]
    store.write("entities", entities)
    store.write("evidence", evidence)
    store.write("relations", relations)

    assert suggest_alias_candidates(store) >= 2
    rows = store.read("alias_candidates", AliasCandidate)
    postal = next(
        item
        for item in rows
        if set(item.compared_entity_ids) == {"entity.postal.camel", "entity.postal.snake"}
    )
    assert "name_token_overlap" in postal.reason or "embedding_similarity" in postal.reason
    zip_pair = next(
        item
        for item in rows
        if set(item.compared_entity_ids) == {"entity.zip.form", "entity.zip.api"}
    )
    assert "relation_similarity" in zip_pair.reason
    assert "same_evidence_neighborhood" in zip_pair.reason


def test_verifier_downgrades_must_when_before_value_or_property_mismatch(tmp_path: Path) -> None:
    store = LocalStore(tmp_path / ".specimpact")
    store.init()
    evidence = Evidence(
        evidence_id="ev.limit",
        document_id="doc",
        section_id="sec",
        chunk_id="chunk",
        quote="requestedCreditLimit upper bound is 999",
        evidence_type="test",
        supports=[EvidenceSupport(type="relation", id="rel.limit")],
        source_location=SourceLocation(file="spec.xlsx", line_start=1, line_end=1),
    )
    relation = Relation(
        relation_id="rel.limit",
        relation_type="VALIDATES",
        source_id="validation.credit_limit",
        target_id="entity.credit_limit",
        evidence_ids=["ev.limit"],
        polarity="explicit",
    )
    store.write("evidence", [evidence])
    priority, _reason = classify_impact(
        store,
        [relation],
        ["ev.limit"],
        ["requestedCreditLimit"],
        "999",
        change_property="max_value",
        artifact_type="ValidationRule",
    )
    assert priority == "must_review"
    priority, _reason = classify_impact(
        store,
        [relation],
        ["ev.limit"],
        ["requestedCreditLimit"],
        "1000",
        change_property="max_value",
        artifact_type="ValidationRule",
    )
    assert priority == "should_review"
    priority, _reason = classify_impact(
        store,
        [relation],
        ["ev.limit"],
        ["requestedCreditLimit"],
        "999",
        change_property="max_value",
        artifact_type="DeploymentJob",
    )
    assert priority == "should_review"


def test_llm_impact_hypothesis_adds_actions_and_reason(tmp_path: Path) -> None:
    store = LocalStore(tmp_path / ".specimpact")
    store.init()
    artifact = Artifact(
        artifact_id="validation.credit_limit",
        artifact_type="ValidationRule",
        display_name="Credit limit upper bound",
    )
    evidence = Evidence(
        evidence_id="ev.limit",
        document_id="doc",
        section_id="sec",
        chunk_id="chunk",
        quote="requestedCreditLimit upper bound is 999",
        evidence_type="test",
        supports=[EvidenceSupport(type="relation", id="rel.limit")],
        source_location=SourceLocation(file="spec.xlsx", line_start=1, line_end=1),
    )
    relation = Relation(
        relation_id="rel.limit",
        relation_type="VALIDATES",
        source_id="validation.credit_limit",
        target_id="entity.credit_limit",
        evidence_ids=["ev.limit"],
        polarity="explicit",
    )
    store.write("artifacts", [artifact])
    store.write("evidence", [evidence])
    atom = [
        ChangeAtom(
            atom_id="atom.limit",
            change_id="change.limit",
            target_terms=["requestedCreditLimit"],
            operation="change_constraint",
            property="max_value",
            before="999",
            after="9999",
        )
    ]
    client = RecordingFakeLLMClient(
        {
            "impact_hypothesis": {
                "impact_type": "boundary_value_change",
                "required_actions": ["Update upper-bound validation and boundary tests."],
                "warnings": ["Check API contract."],
                "uncertainty": "low",
                "reason": "The supplied evidence contains the changed upper bound.",
                "evidence_ids": ["ev.limit"],
                "review_priority_suggestion": "may_review",
            }
        }
    )
    impacts = build_impact_hypotheses(
        store,
        atom,
        [
            RetrievedPath(
                node_id=artifact.artifact_id,
                relations=[relation],
                evidence_ids=["ev.limit"],
            )
        ],
        use_llm=True,
        llm_client=client,
    )
    assert impacts[0].impact_type == "boundary_value_change"
    assert "boundary tests" in impacts[0].required_actions[0]
    assert impacts[0].review_priority == "may_review"
    payload = client.calls[0]["payload"]
    assert "candidate_subgraph" in payload
    assert payload["candidate_subgraph"]["relations"]


@pytest.mark.parametrize(
    ("golden_name", "artifact_id", "change_property"),
    [
        ("利用限度額上限変更.expected.json", "validation.credit_limit_upper_bound", "max_value"),
        ("phone_number_length_change.expected.json", "validation.phone_number_format", "length"),
        (
            "identity_verification_method_change.expected.json",
            "external_if.identity_verification",
            "method",
        ),
        ("external_if_item_added.expected.json", "external_if.credit_screening", "field_added"),
    ],
)
def test_llm_impact_hypothesis_matches_dirty_excel_required_action_goldens(
    tmp_path: Path,
    golden_name: str,
    artifact_id: str,
    change_property: str,
) -> None:
    expected = json.loads(
        (ROOT / "examples" / "dirty_sier_excel" / "goldens" / golden_name).read_text(
            encoding="utf-8"
        )
    )
    expected_actions = expected["expected_required_actions"][artifact_id]
    store = LocalStore(tmp_path / ".specimpact")
    store.init()
    artifact = Artifact(
        artifact_id=artifact_id,
        artifact_type=_artifact_type_for_id(artifact_id),
        display_name=artifact_id,
    )
    evidence = Evidence(
        evidence_id=f"ev.{artifact_id}",
        document_id="doc",
        section_id="sec",
        chunk_id="chunk",
        quote=f"{artifact_id} evidence for {change_property}",
        evidence_type="test",
        supports=[EvidenceSupport(type="relation", id=f"rel.{artifact_id}")],
        source_location=SourceLocation(file="dirty.xlsx", line_start=1, line_end=1),
    )
    relation = Relation(
        relation_id=f"rel.{artifact_id}",
        relation_type="VALIDATES" if artifact.artifact_type == "ValidationRule" else "CALLS",
        source_id=artifact_id,
        target_id="entity.changed",
        evidence_ids=[evidence.evidence_id],
        polarity="explicit",
    )
    store.write("artifacts", [artifact])
    store.write("evidence", [evidence])
    client = RecordingFakeLLMClient(
        {
            "impact_hypothesis": {
                "impact_type": change_property,
                "required_actions": expected_actions,
                "warnings": [],
                "uncertainty": "low",
                "reason": "matches dirty Excel golden",
                "evidence_ids": [evidence.evidence_id],
                "review_priority_suggestion": "should_review",
            }
        }
    )

    impacts = build_impact_hypotheses(
        store,
        [
            ChangeAtom(
                atom_id=f"atom.{artifact_id}",
                change_id=f"change.{artifact_id}",
                target_terms=["entity.changed"],
                operation="change",
                property=change_property,
            )
        ],
        [
            RetrievedPath(
                node_id=artifact_id,
                relations=[relation],
                evidence_ids=[evidence.evidence_id],
            )
        ],
        use_llm=True,
        llm_client=client,
    )

    assert impacts[0].required_actions == expected_actions
    assert client.calls[0]["payload"]["candidate_subgraph"]["relations"]


def test_llm_first_impact_from_dirty_excel_credit_limit(tmp_path: Path) -> None:
    store = LocalStore(tmp_path / ".specimpact")
    ingest_dirty_excel(store, SIER / "docs", SIER / "aliases.yml")
    parse_change_atoms(store, SIER / "changes" / "利用限度額_上限変更.md")

    report = analyze_change_llm_first(store, SIER / "changes" / "利用限度額_上限変更.md")

    must_ids = {
        item.artifact_id
        for item in report.impacts
        if item.review_priority == "must_review"
    }
    assert {
        "screen.enrollment_application",
        "api.enrollment_application",
        "column.requested_credit_limit",
        "validation.credit_limit_upper_bound",
        "external_if.credit_screening",
        "test.credit_limit_boundary",
    } <= must_ids


def test_gui_dirty_excel_and_impact_decision_services(tmp_path: Path) -> None:
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    project = ProjectRegistry(tmp_path / "registry").add(workspace)
    execute(project, "init", {})
    execute(
        project,
        "ingest_dirty_excel",
        {"path": str(SIER / "docs"), "aliases": str(SIER / "aliases.yml")},
    )
    dirty = dirty_excel_data(project)
    assert dirty["summary"]["regions"] >= 6

    result = execute(
        project,
        "analyze_llm_first",
        {"path": str(SIER / "changes" / "利用限度額_上限変更.md")},
    )
    assert result["result"]["candidates"] >= 6
    decision = set_impact_status(
        LocalStore(workspace / ".specimpact"),
        "impact.change.利用限度額_上限変更.api.enrollment_application",
        "accepted",
        "reviewed",
    )
    assert decision.status == "accepted"
    assert impact_decisions_data(project)


def _write_dirty_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "画面項目定義"
    sheet.merge_cells("A1:B1")
    sheet["A1"] = "改訂履歴"
    sheet["A1"].font = Font(bold=True)
    sheet["A1"].fill = PatternFill("solid", fgColor="FFFFCC")
    sheet["A1"].comment = Comment("owner note", "qa")
    sheet["A1"].hyperlink = "https://example.com/spec"
    sheet["A2"] = "1.0"
    sheet["B2"] = "初版"
    sheet.row_dimensions[2].hidden = True
    sheet.column_dimensions["F"].hidden = True
    sheet.append([])
    sheet.append(["画面ID", "画面名", "項目ID", "項目名", "物理名", "入力可否"])
    sheet.append(["SCR-001", "入会申込画面", "F-001", "利用限度額", "requestedCreditLimit", "可"])
    sheet.append([])
    sheet.append(["チェックID", "チェック名", "対象項目", "物理名", "上限"])
    sheet.append(
        ["CHK-001", "利用限度額入力チェック", "利用限度額", "requestedCreditLimit", "999万円"]
    )
    chart = BarChart()
    chart.add_data(Reference(sheet, min_col=1, min_row=2, max_row=2))
    sheet.add_chart(chart, "H2")
    workbook.save(path)


def _artifact_type_for_id(artifact_id: str) -> str:
    if artifact_id.startswith("validation."):
        return "ValidationRule"
    if artifact_id.startswith("external_if."):
        return "ExternalIF"
    if artifact_id.startswith("api."):
        return "API"
    if artifact_id.startswith("test."):
        return "TestCase"
    if artifact_id.startswith("column."):
        return "DBColumn"
    return "Artifact"
