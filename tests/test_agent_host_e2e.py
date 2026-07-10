from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from specimpact.application import (
    ApplicationService,
    HostContext,
    project_from_path,
)
from specimpact.application.approval import ApprovalManager
from specimpact.application.host_workflow import HostWorkflow
from specimpact.dirty_excel.ingestion import ingest_dirty_excel
from specimpact.store import LocalStore

ROOT = Path(__file__).parents[1]
BENCHMARK = ROOT / "examples" / "dirty_sier_excel"
EXTERNAL_IF_WORKBOOK = BENCHMARK / "docs" / "06_external_if_definition.xlsx"


def test_external_if_workbook_uses_dirty_sier_revision_and_mapping_layout() -> None:
    workbook = load_workbook(EXTERNAL_IF_WORKBOOK, data_only=True)
    worksheet = workbook.active

    assert worksheet.title == "外部IF項目定義"
    assert "A1:F1" in {str(item) for item in worksheet.merged_cells.ranges}
    assert [worksheet.cell(4, column).value for column in range(1, 8)] == [
        "IF名",
        "項目名",
        "物理名",
        "型",
        "桁数",
        "必須",
        "備考",
    ]
    assert [worksheet.cell(5, column).value for column in range(1, 8)] == [
        "信用照会連携IF",
        "利用限度額",
        "REQUESTED_CREDIT_LIMIT",
        "NUMBER",
        "7",
        "必須",
        "与信照会依頼に送信",
    ]
    assert [worksheet["A2"].value, worksheet["B2"].value] == ["1.0", "初版"]
    assert worksheet["A1"].fill.fgColor.rgb == "00FFFFCC"
    assert worksheet["A4"].font.bold


def test_agent_host_dirty_excel_credit_limit_e2e_persists_review_and_exports_obsidian(
    tmp_path: Path,
) -> None:
    project_path = tmp_path / "credit-limit-project"
    project_path.mkdir()
    project = project_from_path(project_path)
    store = LocalStore(project_path / ".specimpact")
    summary = ingest_dirty_excel(store, BENCHMARK / "docs", BENCHMARK / "aliases.yml")

    assert summary.workbooks == 6
    workflow = HostWorkflow(
        project,
        HostContext(
            host="e2e-external-host",
            workspace_root=str(project_path),
            project_id=project.project_id,
            model="e2e-host-model",
            capabilities=["sampling"],
            external=True,
        ),
    )
    change_id = _submit_credit_limit_change(workflow)

    prepared_impacts = workflow.prepare_impact_context(change_id)
    assert prepared_impacts.transmission_preview is not None
    assert prepared_impacts.payload["withheld"] is True
    _authorize(
        workflow,
        prepared_impacts.context_id,
        prepared_impacts.transmission_preview.preview_id,
    )

    candidates = workflow._context_record(prepared_impacts.context_id)["context"]["payload"][
        "candidates"
    ]
    selected = _candidates_by_artifact_type(candidates)
    assert set(selected) == {
        "Screen",
        "ValidationRule",
        "API",
        "Column",
        "ExternalIF",
        "TestCase",
    }
    assert all(
        candidate["evidence_ids"] and candidate["relation_ids"]
        for candidate in selected.values()
    )

    submitted = workflow.submit_impact_hypotheses(
        prepared_impacts.context_id,
        {
            "change_id": change_id,
            "hypotheses": [
                {
                    "candidate_node_id": candidate["candidate_node_id"],
                    "atom_id": "atom.credit-limit-upper-bound",
                    "impact_type": "max_value_change",
                    "required_actions": ["上限値と境界値テストを更新する。"],
                    "warnings": [],
                    "uncertainty": "low",
                    "reason": "設計書の明示的な関連とEvidenceを確認した。",
                    "evidence_ids": candidate["evidence_ids"],
                    "relation_ids": candidate["relation_ids"],
                    "review_priority_suggestion": "should_review",
                }
                for candidate in selected.values()
            ],
        },
        "credit-limit-impact-submission",
    )

    assert {item["artifact_type"] for item in submitted.impacts} == set(selected)
    assert all(item["evidence_ids"] and item["relation_paths"] for item in submitted.impacts)

    service = ApplicationService(project)
    accepted_impact = next(
        item for item in submitted.impacts if item["artifact_type"] == "ExternalIF"
    )
    impact_id = f"impact.{change_id}.{accepted_impact['artifact_id']}"
    service.mutate(
        "impact_status",
        {
            "impact_id": impact_id,
            "status": "accepted",
            "reason": "外部IFの上限変更対応を承認した。",
        },
        idempotency_key="accept-external-if-impact",
    )

    decisions = service.impacts(change_id)
    decision = next(item for item in decisions if item["impact_id"] == impact_id)
    assert decision["status"] == "accepted"
    assert decision["candidate_node_id"] == accepted_impact["artifact_id"]
    assert decision["evidence_ids"] == accepted_impact["evidence_ids"]

    vault = project_path / "vault"
    service.mutate(
        "obsidian_export",
        {"path": str(vault)},
        idempotency_key="export-credit-limit-obsidian",
    )
    impact_note = next(
        path
        for path in (vault / "SpecImpact" / "Impacts").glob("*.md")
        if impact_id in path.read_text(encoding="utf-8")
    )
    note_body = impact_note.read_text(encoding="utf-8")
    assert 'status: "accepted"' in note_body
    assert accepted_impact["artifact_id"] in note_body
    assert accepted_impact["evidence_ids"][0] in note_body


def _submit_credit_limit_change(workflow: HostWorkflow) -> str:
    prepared = workflow.prepare_change(
        "# 利用限度額上限変更\n\n"
        "利用限度額の上限を999万円から9999万円に変更する。"
        "requestedCreditLimitとREQUESTED_CREDIT_LIMITを対象とする。"
    )
    assert prepared.transmission_preview is not None
    assert prepared.payload["withheld"] is True
    _authorize(workflow, prepared.context_id, prepared.transmission_preview.preview_id)
    change_id = workflow._context_record(prepared.context_id)["change_id"]
    session = workflow.submit_change_atoms(
        prepared.context_id,
        {
            "change_id": change_id,
            "change_atoms": [
                {
                    "atom_id": "atom.credit-limit-upper-bound",
                    "change_id": change_id,
                    "target_terms": [
                        "利用限度額",
                        "requestedCreditLimit",
                        "REQUESTED_CREDIT_LIMIT",
                    ],
                    "operation": "change_constraint",
                    "property": "max_value",
                    "before": "999",
                    "after": "9999",
                    "likely_node_types": [
                        "Screen",
                        "ValidationRule",
                        "API",
                        "DBColumn",
                        "ExternalIF",
                        "TestCase",
                    ],
                }
            ],
        },
        "credit-limit-atom-submission",
    )
    assert session.status == "atoms_ready"
    return change_id


def _authorize(workflow: HostWorkflow, context_id: str, preview_id: str) -> None:
    grant = ApprovalManager(workflow.project).issue_grant(preview_id, decision="approve")
    assert workflow.authorize_context(context_id, grant.token).context_id == context_id


def _candidates_by_artifact_type(
    candidates: list[dict[str, object]],
) -> dict[str, dict[str, object]]:
    required_types = {"Screen", "ValidationRule", "API", "Column", "ExternalIF", "TestCase"}
    return {
        str(candidate["artifact"]["artifact_type"]): candidate
        for candidate in candidates
        if str(candidate["artifact"]["artifact_type"]) in required_types
    }
