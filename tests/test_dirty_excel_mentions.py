from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from specimpact.dirty_excel.ingestion import ingest_dirty_excel
from specimpact.dirty_excel.models import DirtyRegion
from specimpact.impact_management.review_session import analyze_change_llm_first
from specimpact.models import Evidence, Relation
from specimpact.store import LocalStore


def test_sheet_mention_graph_connects_alias_to_cell_evidence(tmp_path: Path) -> None:
    docs = tmp_path / "docs"
    docs.mkdir()
    workbook_path = docs / "テーブル定義書.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "プロジェクト"
    sheet.append(["論理名称", "物理名称", "データ型", "桁数"])
    sheet.append(["プロジェクト名", "PROJECT_NAME", "VARCHAR", 128])
    workbook.save(workbook_path)
    aliases = tmp_path / "aliases.yml"
    aliases.write_text(
        "aliases:\n"
        "  entity.project_name:\n"
        "    canonical_type: BusinessField\n"
        "    aliases: [プロジェクト名, PROJECT_NAME, projectName]\n",
        encoding="utf-8",
    )
    change = tmp_path / "change.md"
    change.write_text(
        "# プロジェクト名の最大長変更\n\n"
        "プロジェクト名の最大長を128文字から256文字へ変更する。\n",
        encoding="utf-8",
    )
    store = LocalStore(tmp_path / ".specimpact")

    ingest_dirty_excel(store, docs, aliases)

    relations = [
        relation
        for relation in store.read("relations", Relation)
        if relation.target_id == "entity.project_name"
    ]
    assert any(relation.relation_type == "DEFINES" for relation in relations)
    evidence = {item.evidence_id: item for item in store.read("evidence", Evidence)}
    mention = next(
        evidence[evidence_id]
        for relation in relations
        for evidence_id in relation.evidence_ids
        if evidence[evidence_id].evidence_type == "dirty_excel_cell_mention"
    )
    assert "[A2] プロジェクト名" in mention.quote
    assert "[D2] 128" in mention.quote
    assert mention.source_location.file.endswith("テーブル定義書.xlsx")

    report = analyze_change_llm_first(store, change, no_llm=True)
    assert any(
        impact.artifact_type == "Table" and "プロジェクト" in impact.display_name
        for impact in report.impacts
    )


def test_sheet_mention_evidence_is_available_to_region_highlighting(tmp_path: Path) -> None:
    workbook_path = tmp_path / "画面設計書.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "画面項目"
    sheet.append(["項目名", "物理名", "桁数"])
    sheet.append(["プロジェクト名", "projectName", 128])
    workbook.save(workbook_path)
    aliases = tmp_path / "aliases.yml"
    aliases.write_text(
        "aliases:\n"
        "  entity.project_name:\n"
        "    canonical_type: BusinessField\n"
        "    aliases: [プロジェクト名, projectName]\n",
        encoding="utf-8",
    )
    store = LocalStore(tmp_path / ".specimpact")

    ingest_dirty_excel(store, workbook_path, aliases)

    relation = next(
        item
        for item in store.read("relations", Relation)
        if item.target_id == "entity.project_name"
        and any(
            evidence.evidence_type == "dirty_excel_cell_mention"
            for evidence in store.read("evidence", Evidence)
            if evidence.evidence_id in item.evidence_ids
        )
    )
    regions = store.read("dirty_regions", DirtyRegion)
    assert any(set(relation.evidence_ids) & set(region.evidence_ids) for region in regions)


def test_sheet_mention_graph_keeps_late_matching_rows_in_large_region(tmp_path: Path) -> None:
    workbook_path = tmp_path / "単体テスト仕様書.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TEST01"
    sheet.append(["No.", "入力条件", "期待結果"])
    sheet.append([1, "プロジェクト名を128文字で入力", "正常"])
    for row in range(2, 18):
        sheet.append([row, "別項目の確認", "正常"])
    sheet.append([18, "プロジェクト名を129文字で入力", "エラー"])
    workbook.save(workbook_path)
    aliases = tmp_path / "aliases.yml"
    aliases.write_text(
        "aliases:\n"
        "  entity.project_name:\n"
        "    canonical_type: BusinessField\n"
        "    aliases: [プロジェクト名]\n",
        encoding="utf-8",
    )
    store = LocalStore(tmp_path / ".specimpact")

    ingest_dirty_excel(store, workbook_path, aliases)

    relation = next(
        item
        for item in store.read("relations", Relation)
        if item.target_id == "entity.project_name" and item.relation_type == "COVERS"
    )
    evidence = {item.evidence_id: item for item in store.read("evidence", Evidence)}
    quotes = [evidence[evidence_id].quote for evidence_id in relation.evidence_ids]
    assert any("[B2] プロジェクト名を128文字で入力" in quote for quote in quotes)
    assert any("[B19] プロジェクト名を129文字で入力" in quote for quote in quotes)


def test_mixed_japanese_artifact_names_do_not_collide(tmp_path: Path) -> None:
    docs = tmp_path / "docs"
    docs.mkdir()
    for filename in ("メッセージ設計書(画面).xlsx", "メッセージ設計書(バッチ).xlsx"):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "ja"
        sheet.append(["ID", "メッセージ"])
        sheet.append(["length", "プロジェクト名は128文字以下で入力してください"])
        workbook.save(docs / filename)
    aliases = tmp_path / "aliases.yml"
    aliases.write_text(
        "aliases:\n"
        "  entity.project_name:\n"
        "    canonical_type: BusinessField\n"
        "    aliases: [プロジェクト名]\n",
        encoding="utf-8",
    )
    store = LocalStore(tmp_path / ".specimpact")

    ingest_dirty_excel(store, docs, aliases)

    message_relations = [
        relation
        for relation in store.read("relations", Relation)
        if relation.target_id == "entity.project_name"
        and relation.relation_type == "VALIDATES"
    ]
    assert len({relation.source_id for relation in message_relations}) == 2
