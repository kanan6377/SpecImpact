from __future__ import annotations

import csv
import re
from collections import Counter
from pathlib import Path
from typing import Any
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from specimpact.extraction import (
    AliasCatalog,
    GraphRecords,
    artifact_for,
    entity_for,
    make_document,
    relation_with_evidence,
)
from specimpact.models import Artifact, Relation
from specimpact.store import LocalStore


def ingest_csv(store: LocalStore, path: Path) -> list[dict[str, Any]]:
    if not path.is_file():
        raise ValueError(f"CSV source does not exist: {path}")
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise ValueError("CSV header row is required")
        rows = list(reader)
    return _ingest_tables(store, path, [(path.stem, reader.fieldnames, rows)], "csv")


def ingest_excel(
    store: LocalStore,
    path: Path,
    aliases_path: Path | None = None,
    profile: str | None = None,
) -> list[dict[str, Any]]:
    store.init()
    if aliases_path:
        if not aliases_path.is_file():
            raise ValueError(f"Aliases file does not exist: {aliases_path}")
        alias_text = aliases_path.read_text(encoding="utf-8")
        AliasCatalog.parse(alias_text, aliases_path)
        store.write_text(store.root / "aliases.yml", alias_text)
    if path.is_dir():
        workbooks = sorted(item for item in path.iterdir() if item.suffix.lower() == ".xlsx")
        if not workbooks:
            raise ValueError(f"Excel directory contains no .xlsx files: {path}")
        records: list[dict[str, Any]] = []
        health = inspect_excel_folder(path)
        for workbook in workbooks:
            records.extend(
                _ingest_excel_workbook(
                    store,
                    workbook,
                    source_key=workbook.name,
                    force_sier=profile == "sier",
                )
            )
        health["detected_artifacts"] = len(store.read("artifacts", Artifact))
        health["possible_relations"] = len(store.read("relations", Relation))
        store.write_json(store.root / "health_check.json", health)
        return records
    return _ingest_excel_workbook(store, path, source_key=path.name, force_sier=profile == "sier")


def inspect_excel_folder(path: Path) -> dict[str, Any]:
    if path.is_file():
        workbooks = [path]
    elif path.is_dir():
        workbooks = sorted(item for item in path.iterdir() if item.suffix.lower() == ".xlsx")
    else:
        raise ValueError(f"Excel source does not exist: {path}")
    if not workbooks:
        raise ValueError(f"Excel directory contains no .xlsx files: {path}")
    health: dict[str, Any] = {
        "workbooks": len(workbooks),
        "sheets": 0,
        "detected_artifacts": 0,
        "possible_relations": 0,
        "merged_cells": 0,
        "hidden_sheets": 0,
        "textboxes_or_shapes": 0,
        "multi_header_candidate_sheets": [],
        "revision_history_sheets": [],
        "duplicate_field_names": [],
        "alias_candidates": [],
        "workbook_details": [],
        "warnings": [],
    }
    field_names: Counter[str] = Counter()
    alias_groups: dict[str, set[str]] = {}
    for workbook_path in workbooks:
        try:
            workbook = load_workbook(workbook_path, read_only=False, data_only=True)
        except (
            BadZipFile,
            InvalidFileException,
            KeyError,
            OSError,
            ParseError,
            ValueError,
        ) as error:
            raise ValueError(f"Invalid Excel source: {workbook_path}") from error
        details = {"file": workbook_path.name, "sheets": []}
        for sheet in workbook.worksheets:
            health["sheets"] += 1
            rows = list(sheet.iter_rows(values_only=True))
            headers = [str(value) for value in rows[0] if value not in (None, "")] if rows else []
            header_rows = [
                index
                for index, row in enumerate(rows[:8], start=1)
                if sum(1 for value in row if value not in (None, "")) >= 3
            ]
            if len(header_rows) > 1:
                health["multi_header_candidate_sheets"].append(
                    f"{workbook_path.name} / {sheet.title}"
                )
            if sheet.sheet_state != "visible":
                health["hidden_sheets"] += 1
            if "改訂" in sheet.title or "履歴" in sheet.title:
                health["revision_history_sheets"].append(f"{workbook_path.name} / {sheet.title}")
            merged_count = len(sheet.merged_cells.ranges)
            health["merged_cells"] += merged_count
            rows_with_values = sum(1 for row in rows if any(value is not None for value in row))
            data_rows = max(0, rows_with_values - 1)
            health["detected_artifacts"] += data_rows
            health["possible_relations"] += data_rows
            for row in rows[1:]:
                values = [str(value) for value in row if value not in (None, "")]
                for value in values:
                    if _looks_like_field(value):
                        field_names[value] += 1
                _collect_alias_candidates(values, alias_groups)
            details["sheets"].append(
                {
                    "name": sheet.title,
                    "rows": rows_with_values,
                    "columns": len(headers),
                    "merged_cells": merged_count,
                    "estimated_type": _estimate_sheet_type(
                        workbook_path.name,
                        sheet.title,
                        headers,
                    ),
                }
            )
        health["workbook_details"].append(details)
    health["duplicate_field_names"] = [
        {"name": name, "count": count} for name, count in sorted(field_names.items()) if count > 1
    ]
    health["alias_candidates"] = [
        sorted(values) for values in alias_groups.values() if len(values) > 1
    ]
    if health["merged_cells"]:
        health["warnings"].append("結合セルがあります。必要に応じて表形式へ整形してください。")
    if health["multi_header_candidate_sheets"]:
        health["warnings"].append("複数のヘッダー候補があるシートがあります。")
    return health


def _ingest_excel_workbook(
    store: LocalStore,
    path: Path,
    *,
    source_key: str,
    force_sier: bool = False,
) -> list[dict[str, Any]]:
    if not path.is_file():
        raise ValueError(f"Excel source does not exist: {path}")
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, KeyError, OSError, ParseError, ValueError) as error:
        raise ValueError(f"Invalid Excel source: {path}") from error
    sheet_text = _workbook_text(workbook)
    aliases = AliasCatalog.load(store.root / "aliases.yml")
    if force_sier or _looks_like_sier_workbook(path.name, workbook):
        return _ingest_sier_workbook(store, path, workbook, aliases, sheet_text, source_key)
    tables = []
    for sheet in workbook.worksheets:
        values = list(sheet.iter_rows(values_only=True))
        if not values or not any(values[0]):
            raise ValueError(f"Excel header row is required: {sheet.title}")
        headers = [str(value) if value is not None else "" for value in values[0]]
        rows = [
            {headers[index]: value for index, value in enumerate(row) if headers[index]}
            for row in values[1:]
            if any(value is not None for value in row)
        ]
        tables.append((sheet.title, headers, rows))
    return _ingest_tables(
        store,
        path,
        tables,
        "excel",
        text=sheet_text,
        source_key=source_key,
    )


def _ingest_tables(
    store: LocalStore,
    path: Path,
    tables: list[tuple[str, list[str], list[dict[str, Any]]]],
    source_type: str,
    *,
    text: str | None = None,
    source_key: str | None = None,
) -> list[dict[str, Any]]:
    store.init()
    aliases = AliasCatalog.load(store.root / "aliases.yml")
    document, section, chunk = make_document(
        path,
        source_type,
        text=text or path.name,
        source_key=source_key or path.name,
    )
    graph = GraphRecords(documents=[document], sections=[section], chunks=[chunk])
    records = []
    for table_name, headers, rows in tables:
        table = artifact_for(table_name, "Table", document.document_id, aliases)
        graph.artifacts.append(table)
        records.append(
            {
                "artifact_id": table.artifact_id,
                "artifact_type": "Table",
                "display_name": table_name,
                "headers": headers,
                "rows": rows,
                "source": path.as_posix(),
            }
        )
        for header in headers:
            column = artifact_for(f"{table_name}.{header}", "Column", document.document_id, aliases)
            entity = entity_for(header, document.document_id, aliases)
            graph.artifacts.append(column)
            graph.entities.append(entity)
            relation, evidence = relation_with_evidence(
                source_id=column.artifact_id,
                target_id=entity.entity_id,
                relation_type="DEFINES",
                document=document,
                section_id=section.section_id,
                chunk_id=chunk.chunk_id,
                line_number=1,
                quote=header,
            )
            graph.relations.append(relation)
            graph.evidence.append(evidence)
    store.merge_graph(**graph.__dict__)
    return records


def _ingest_sier_workbook(
    store: LocalStore,
    path: Path,
    workbook,
    aliases: AliasCatalog,
    sheet_text: str,
    source_key: str,
) -> list[dict[str, Any]]:
    document, section, chunk = make_document(path, "excel", text=sheet_text, source_key=source_key)
    graph = GraphRecords(documents=[document], sections=[section], chunks=[chunk])
    records: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        values = list(sheet.iter_rows(values_only=True))
        if not values or not any(values[0]):
            continue
        headers = [str(value).strip() if value is not None else "" for value in values[0]]
        header_columns = {
            header: index + 1
            for index, header in enumerate(headers)
            if header
        }
        rows = [
            {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
            for row in values[1:]
            if any(value is not None for value in row)
        ]
        kind = _estimate_sheet_type(path.name, sheet.title, headers)
        for row_number, row in enumerate(rows, start=2):
            args = (
                graph,
                document,
                section.section_id,
                chunk.chunk_id,
                sheet.title,
                row_number,
                row,
                header_columns,
                aliases,
            )
            if kind == "screen":
                _add_screen_row(*args)
            elif kind == "api":
                _add_api_row(*args)
            elif kind == "table":
                _add_table_row(*args)
            elif kind == "validation":
                _add_validation_row(*args)
            elif kind == "external_if":
                _add_external_if_row(*args)
            elif kind == "test":
                _add_test_row(*args)
        records.append(
            {
                "artifact_type": kind,
                "display_name": sheet.title,
                "headers": headers,
                "rows": rows,
                "source": path.as_posix(),
            }
        )
    store.merge_graph(**graph.__dict__)
    return records


def _add_screen_row(
    graph,
    document,
    section_id,
    chunk_id,
    sheet,
    row_number,
    row,
    header_columns,
    aliases,
) -> None:
    screen_name = _cell(row, "画面名") or _cell(row, "screen_name")
    item_name = _cell(row, "項目名") or _cell(row, "item_name")
    physical_name = _cell(row, "物理名") or _cell(row, "physical_name")
    api_name = _cell(row, "呼び出すAPI") or _cell(row, "api")
    input_flag = _cell(row, "入力可否")
    relation_type = "DISPLAYS" if "不可" in input_flag else "DEFINES"
    if not screen_name or not (item_name or physical_name):
        return
    screen = _append_artifact(graph, screen_name, "Screen", document.document_id, aliases)
    for value in [item_name, physical_name]:
        if value:
            header = "項目名" if value == item_name else "物理名"
            _append_entity_relation(
                graph, screen.artifact_id, value, relation_type, document, section_id, chunk_id,
                sheet, row_number, _row_quote(row), aliases, "screen_field_definition",
                header_columns, header
            )
    if api_name:
        api = _append_artifact(graph, api_name, "API", document.document_id, aliases)
        _append_artifact_relation(
            graph, screen.artifact_id, api.artifact_id, "CALLS", document, section_id,
            chunk_id, sheet, row_number, _row_quote(row), "screen_api_call",
            header_columns, "呼び出すAPI"
        )


def _add_api_row(
    graph,
    document,
    section_id,
    chunk_id,
    sheet,
    row_number,
    row,
    header_columns,
    aliases,
) -> None:
    api_name = _cell(row, "API名") or _cell(row, "api_name")
    field_name = _cell(row, "物理名") or _cell(row, "項目名") or _cell(row, "field")
    direction = str(_cell(row, "区分") or _cell(row, "direction") or "request").lower()
    if not api_name or not field_name:
        return
    api = _append_artifact(graph, api_name, "API", document.document_id, aliases)
    relation_type = (
        "RESPONSE_FIELD"
        if "response" in direction or "レスポンス" in direction
        else "REQUEST_FIELD"
    )
    header = "物理名" if _cell(row, "物理名") else "項目名"
    _append_entity_relation(
        graph, api.artifact_id, field_name, relation_type, document, section_id, chunk_id,
        sheet, row_number, _row_quote(row), aliases, "api_request_definition",
        header_columns, header
    )


def _add_table_row(
    graph,
    document,
    section_id,
    chunk_id,
    sheet,
    row_number,
    row,
    header_columns,
    aliases,
) -> None:
    table_name = _cell(row, "テーブル名") or _cell(row, "table_name") or sheet
    column_name = _cell(row, "カラム名") or _cell(row, "column_name")
    logical_name = _cell(row, "カラム論理名") or _cell(row, "logical_name")
    if not table_name or not column_name:
        return
    _append_artifact(graph, table_name, "Table", document.document_id, aliases)
    column = _append_artifact(
        graph, f"{table_name}.{column_name}", "Column", document.document_id, aliases
    )
    for value in [column_name, logical_name]:
        if value:
            header = "カラム名" if value == column_name else "カラム論理名"
            _append_entity_relation(
                graph, column.artifact_id, value, "DEFINES", document, section_id,
                chunk_id, sheet, row_number, _row_quote(row), aliases, "db_column_definition",
                header_columns, header
            )


def _add_validation_row(
    graph,
    document,
    section_id,
    chunk_id,
    sheet,
    row_number,
    row,
    header_columns,
    aliases,
) -> None:
    name = _cell(row, "チェック名") or _cell(row, "チェックID")
    target = _cell(row, "対象項目") or _cell(row, "物理名")
    if not name or not target:
        return
    validation = _append_artifact(graph, name, "ValidationRule", document.document_id, aliases)
    header = "対象項目" if _cell(row, "対象項目") else "物理名"
    _append_entity_relation(
        graph, validation.artifact_id, target, "VALIDATES", document, section_id, chunk_id,
        sheet, row_number, _row_quote(row), aliases, "validation_rule_definition",
        header_columns, header
    )


def _add_external_if_row(
    graph,
    document,
    section_id,
    chunk_id,
    sheet,
    row_number,
    row,
    header_columns,
    aliases,
) -> None:
    if_name = _cell(row, "IF名") or _cell(row, "if_name")
    field_name = _cell(row, "物理名") or _cell(row, "項目名")
    if not if_name or not field_name:
        return
    external_if = _append_artifact(graph, if_name, "ExternalIF", document.document_id, aliases)
    header = "物理名" if _cell(row, "物理名") else "項目名"
    _append_entity_relation(
        graph, external_if.artifact_id, field_name, "SENDS", document, section_id, chunk_id,
        sheet, row_number, _row_quote(row), aliases, "external_mapping_definition",
        header_columns, header
    )


def _add_test_row(
    graph,
    document,
    section_id,
    chunk_id,
    sheet,
    row_number,
    row,
    header_columns,
    aliases,
) -> None:
    test_name = _cell(row, "テスト名") or _cell(row, "テストケースID")
    target = _cell(row, "対象項目") or _cell(row, "確認観点")
    if not test_name or not target:
        return
    test = _append_artifact(graph, test_name, "TestCase", document.document_id, aliases)
    header = "対象項目" if _cell(row, "対象項目") else "確認観点"
    _append_entity_relation(
        graph, test.artifact_id, target, "VALIDATES", document, section_id, chunk_id,
        sheet, row_number, _row_quote(row), aliases, "test_coverage_definition",
        header_columns, header
    )


def _append_artifact(
    graph: GraphRecords,
    name: str,
    item_type: str,
    document_id: str,
    aliases: AliasCatalog,
):
    artifact = artifact_for(str(name), item_type, document_id, aliases)
    graph.artifacts.append(artifact)
    return artifact


def _append_entity_relation(
    graph: GraphRecords,
    source_id: str,
    target_name: str,
    relation_type: str,
    document,
    section_id: str,
    chunk_id: str,
    sheet: str,
    row_number: int,
    quote: str,
    aliases: AliasCatalog,
    evidence_type: str,
    header_columns: dict[str, int],
    evidence_header: str,
) -> None:
    entity = entity_for(str(target_name), document.document_id, aliases)
    graph.entities.append(entity)
    relation, evidence = relation_with_evidence(
        source_id=source_id,
        target_id=entity.entity_id,
        relation_type=relation_type,
        document=document,
        section_id=section_id,
        chunk_id=chunk_id,
        line_number=row_number,
        quote=f"[{sheet}!{_best_cell(row_number, header_columns, evidence_header)}] {quote}",
        evidence_type=evidence_type,
    )
    graph.relations.append(relation)
    graph.evidence.append(evidence)


def _append_artifact_relation(
    graph: GraphRecords,
    source_id: str,
    target_id: str,
    relation_type: str,
    document,
    section_id: str,
    chunk_id: str,
    sheet: str,
    row_number: int,
    quote: str,
    evidence_type: str,
    header_columns: dict[str, int],
    evidence_header: str,
) -> None:
    relation, evidence = relation_with_evidence(
        source_id=source_id,
        target_id=target_id,
        relation_type=relation_type,
        document=document,
        section_id=section_id,
        chunk_id=chunk_id,
        line_number=row_number,
        quote=f"[{sheet}!{_best_cell(row_number, header_columns, evidence_header)}] {quote}",
        evidence_type=evidence_type,
        target_support_type="artifact",
    )
    graph.relations.append(relation)
    graph.evidence.append(evidence)


def _workbook_text(workbook) -> str:
    lines = []
    for sheet in workbook.worksheets:
        lines.append(f"# {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            lines.append("\t".join("" if value is None else str(value) for value in row))
    return "\n".join(lines)


def _looks_like_sier_workbook(name: str, workbook) -> bool:
    if any(
        token in name
        for token in ("画面", "API", "テーブル", "入力チェック", "外部IF", "試験")
    ):
        return True
    keywords = {
        "画面ID",
        "画面名",
        "項目名",
        "物理名",
        "API名",
        "endpoint",
        "エンドポイント",
        "テーブル名",
        "カラム名",
        "チェック内容",
        "外部IF",
        "IF名",
        "送信項目",
        "受信項目",
        "テストケースID",
        "期待結果",
    }
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(min_row=1, max_row=3, values_only=True))
        visible_values = {
            str(value).strip()
            for row in rows
            for value in row
            if value not in (None, "")
        }
        if sheet.title in keywords or len(visible_values & keywords) >= 2:
            return True
    return False


def _estimate_sheet_type(file_name: str, sheet_name: str, headers: list[str]) -> str:
    source_name = f"{file_name} {sheet_name}"
    if "画面" in source_name:
        return "screen"
    if "入力チェック" in source_name or "チェック一覧" in source_name:
        return "validation"
    if "外部IF" in source_name:
        return "external_if"
    if "試験" in source_name or "テスト" in source_name:
        return "test"
    if "API" in source_name:
        return "api"
    if "テーブル" in source_name:
        return "table"
    joined = " ".join([file_name, sheet_name, *headers])
    if "入力チェック" in joined or "チェック" in joined:
        return "validation"
    if "外部IF" in joined or "IF名" in joined:
        return "external_if"
    if "試験" in joined or "テスト" in joined:
        return "test"
    if "API" in joined:
        return "api"
    if "テーブル" in joined or "カラム" in joined or "DB" in joined:
        return "table"
    if "画面" in joined:
        return "screen"
    return "table"


def _cell(row: dict[str, Any], name: str) -> str:
    value = row.get(name)
    return "" if value is None else str(value).strip()


def _row_quote(row: dict[str, Any]) -> str:
    return " / ".join(f"{key}={value}" for key, value in row.items() if value not in (None, ""))


def _best_cell(row_number: int, header_columns: dict[str, int], evidence_header: str) -> str:
    column = header_columns.get(evidence_header, 1)
    return f"{get_column_letter(column)}{row_number}"


def _looks_like_field(value: str) -> bool:
    return bool(re.search(r"[A-Za-z_]{4,}|額|金額|ID|コード", value))


def _collect_alias_candidates(values: list[str], groups: dict[str, set[str]]) -> None:
    logical = [value for value in values if any(token in value for token in ("額", "金額", "限度"))]
    physical = [
        value for value in values if any(char.isascii() and char.isalpha() for char in value)
    ]
    for left in logical:
        key = left.replace("利用", "").replace("希望", "")
        groups.setdefault(key, set()).add(left)
        for right in physical:
            groups[key].add(right)
