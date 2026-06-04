from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from specimpact.core import latest_run_dir
from specimpact.models import Evidence
from specimpact.store import LocalStore


def export_report_excel(store: LocalStore, output_path: Path | None = None) -> Path:
    run_dir = latest_run_dir(store)
    report = json.loads((run_dir / "report.json").read_text(encoding="utf-8"))
    evidence = {item.evidence_id: item for item in store.read("evidence", Evidence)}
    health = _read_json(store.root / "health_check.json", {})
    aliases = _read_yaml(store.root / "aliases.yml")
    output = output_path or (run_dir / "report.xlsx")
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    _write_rows(
        summary,
        [
            ["run_id", report["run_id"]],
            ["change", report["change"]["title"]],
            ["must_review", len(report["must_review"])],
            ["should_review", len(report["should_review"])],
            ["may_review", len(report["may_review"])],
            ["note", "このレポートは影響確定結果ではなく、レビュー候補です。"],
        ],
    )
    candidates = workbook.create_sheet("ReviewCandidates")
    candidate_rows = [[
        "priority",
        "artifact_type",
        "display_name",
        "reason",
        "file",
        "sheet",
        "row",
        "cell",
        "relation_path",
        "status",
        "owner",
        "comment",
    ]]
    for priority in ("must_review", "should_review", "may_review"):
        for impact in report[priority]:
            first = _first_evidence(impact, evidence)
            location = _location(first)
            candidate_rows.append(
                [
                    priority,
                    impact["artifact_type"],
                    impact["display_name"],
                    impact["reason"],
                    location["file"],
                    location["sheet"],
                    location["row"],
                    location["cell"],
                    "\n".join(impact.get("relation_paths", [])),
                    "",
                    "",
                    "",
                ]
            )
    _write_rows(candidates, candidate_rows)
    evidence_sheet = workbook.create_sheet("Evidence")
    _write_rows(
        evidence_sheet,
        [["evidence_id", "file", "sheet", "row", "cell", "quote"]]
        + [
            [
                item.evidence_id,
                _location(item)["file"],
                _location(item)["sheet"],
                _location(item)["row"],
                _location(item)["cell"],
                item.quote,
            ]
            for item in evidence.values()
        ],
    )
    alias_sheet = workbook.create_sheet("AliasCandidates")
    alias_rows = [["target_id", "canonical_type", "aliases"]]
    for target_id, details in (aliases.get("aliases") or {}).items():
        alias_rows.append(
            [
                target_id,
                details.get("canonical_type", ""),
                ", ".join(details.get("aliases", [])),
            ]
        )
    _write_rows(alias_sheet, alias_rows)
    health_sheet = workbook.create_sheet("HealthCheck")
    health_rows = [["key", "value"]]
    for key, value in health.items():
        rendered = (
            json.dumps(value, ensure_ascii=False)
            if isinstance(value, (list, dict))
            else value
        )
        health_rows.append([key, rendered])
    _write_rows(health_sheet, health_rows)
    for sheet in workbook.worksheets:
        _format_sheet(sheet)
    workbook.save(output)
    return output


def _write_rows(sheet, rows: list[list[Any]]) -> None:
    for row in rows:
        sheet.append(row)


def _format_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="E2E8F0")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for column in sheet.columns:
        letter = get_column_letter(column[0].column)
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 64)
        sheet.column_dimensions[letter].width = width


def _first_evidence(impact: dict[str, Any], evidence: dict[str, Evidence]) -> Evidence | None:
    return next(
        (evidence[item_id] for item_id in impact.get("evidence_ids", []) if item_id in evidence),
        None,
    )


def _location(evidence: Evidence | None) -> dict[str, Any]:
    if evidence is None:
        return {"file": "", "sheet": "", "row": "", "cell": ""}
    match = re.search(r"\[([^!\]]+)!([A-Z]+\d+)\]", evidence.quote)
    return {
        "file": Path(evidence.source_location.file).name,
        "sheet": match.group(1) if match else "",
        "row": evidence.source_location.line_start,
        "cell": match.group(2) if match else "",
    }


def _read_json(path: Path, default: Any) -> Any:
    return json.loads(path.read_text(encoding="utf-8")) if path.exists() else default


def _read_yaml(path: Path) -> dict[str, Any]:
    if not path.exists() or not path.read_text(encoding="utf-8").strip():
        return {}
    return yaml.safe_load(path.read_text(encoding="utf-8")) or {}
