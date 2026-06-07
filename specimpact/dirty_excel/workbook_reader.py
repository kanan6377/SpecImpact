from __future__ import annotations

import hashlib
import shutil
from pathlib import Path
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from specimpact.dirty_excel.models import CellStyle, DirtyCell, DirtySheet, DirtyWorkbook


def workbook_id_for(path: Path) -> str:
    digest = hashlib.sha1(path.read_bytes()).hexdigest()[:12]
    return f"wb.{_slug(path.stem)}.{digest}"


def read_dirty_workbook(path: Path) -> tuple[DirtyWorkbook, list[DirtySheet], list[DirtyCell]]:
    if not path.is_file():
        raise ValueError(f"Excel source does not exist: {path}")
    try:
        workbook = load_workbook(path, read_only=False, data_only=True)
    except (BadZipFile, InvalidFileException, KeyError, OSError, ParseError, ValueError) as error:
        raise ValueError(f"Invalid Excel source: {path}") from error
    workbook_id = workbook_id_for(path)
    sheets: list[DirtySheet] = []
    cells: list[DirtyCell] = []
    for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
        sheet_id = f"{workbook_id}.sheet.{sheet_index:03d}"
        merged_lookup = _merged_lookup(sheet)
        sheets.append(
            DirtySheet(
                workbook_id=workbook_id,
                sheet_id=sheet_id,
                sheet_name=sheet.title,
                sheet_index=sheet_index,
                max_row=sheet.max_row or 0,
                max_column=sheet.max_column or 0,
                hidden=sheet.sheet_state != "visible",
                image_count=len(getattr(sheet, "_images", []) or []),
                chart_count=len(getattr(sheet, "_charts", []) or []),
                table_count=len(getattr(sheet, "tables", {}) or {}),
                unsupported_drawings=_unsupported_drawings(sheet),
            )
        )
        for row in sheet.iter_rows():
            for cell in row:
                if not _should_keep_cell(cell, merged_lookup):
                    continue
                address = cell.coordinate
                cells.append(
                    DirtyCell(
                        workbook_id=workbook_id,
                        file_path=path.as_posix(),
                        sheet_id=sheet_id,
                        sheet_name=sheet.title,
                        cell=address,
                        evidence_id=f"cell.{_slug(sheet_id)}.{address.lower()}",
                        value=None if cell.value is None else str(cell.value),
                        data_type=str(cell.data_type),
                        row=cell.row,
                        column=cell.column,
                        merged_range=merged_lookup.get(address),
                        style=_style_for(cell),
                        hyperlink=cell.hyperlink.target if cell.hyperlink else None,
                        comment=cell.comment.text if cell.comment else None,
                        is_hidden_row=bool(sheet.row_dimensions[cell.row].hidden),
                        is_hidden_col=bool(
                            sheet.column_dimensions[get_column_letter(cell.column)].hidden
                        ),
                    )
                )
    dirty_workbook = DirtyWorkbook(
        workbook_id=workbook_id,
        file_path=path.as_posix(),
        original_path="",
        normalized_path="",
        sheet_ids=[sheet.sheet_id for sheet in sheets],
        warnings=[
            warning
            for sheet in sheets
            for warning in _drawing_warnings(sheet.sheet_name, sheet.unsupported_drawings)
        ],
    )
    return dirty_workbook, sheets, cells


def preserve_original(path: Path, project_root: Path, workbook_id: str) -> Path:
    original_dir = project_root / "sources" / "original"
    original_dir.mkdir(parents=True, exist_ok=True)
    target = original_dir / f"{workbook_id}{path.suffix.lower()}"
    shutil.copy2(path, target)
    return target


def write_normalized(project_root: Path, workbook: DirtyWorkbook, cells: list[DirtyCell]) -> Path:
    normalized_dir = project_root / "sources" / "normalized"
    normalized_dir.mkdir(parents=True, exist_ok=True)
    target = normalized_dir / f"{workbook.workbook_id}.workbook.jsonl"
    content = "".join(cell.model_dump_json() + "\n" for cell in cells)
    target.write_text(content, encoding="utf-8")
    return target


def _should_keep_cell(cell: Cell, merged_lookup: dict[str, str]) -> bool:
    if cell.value not in (None, ""):
        return True
    if cell.coordinate in merged_lookup:
        return True
    style = _style_for(cell)
    return bool(
        style.fill_color
        or style.font_bold
        or style.border
        or cell.comment
        or cell.hyperlink
    )


def _merged_lookup(sheet) -> dict[str, str]:
    result: dict[str, str] = {}
    for merged in sheet.merged_cells.ranges:
        range_text = str(merged)
        for row in range(merged.min_row, merged.max_row + 1):
            for column in range(merged.min_col, merged.max_col + 1):
                result[sheet.cell(row=row, column=column).coordinate] = range_text
    return result


def _style_for(cell: Cell) -> CellStyle:
    return CellStyle(
        fill_color=_fill_color(cell),
        font_bold=bool(cell.font and cell.font.bold),
        border=_has_border(cell),
        number_format=cell.number_format if cell.number_format else None,
        horizontal=cell.alignment.horizontal,
        vertical=cell.alignment.vertical,
    )


def _fill_color(cell: Cell) -> str | None:
    fill = cell.fill
    if not fill or fill.patternType in (None, "none"):
        return None
    color = fill.fgColor
    if color.type == "rgb" and color.rgb:
        return color.rgb
    if color.type == "indexed" and color.indexed is not None:
        return f"indexed:{color.indexed}"
    if color.type == "theme" and color.theme is not None:
        return f"theme:{color.theme}"
    return None


def _has_border(cell: Cell) -> bool:
    border = cell.border
    return any(
        side and side.style
        for side in (border.left, border.right, border.top, border.bottom)
    )


def _unsupported_drawings(sheet) -> list[str]:
    drawings: list[str] = []
    image_count = len(getattr(sheet, "_images", []) or [])
    chart_count = len(getattr(sheet, "_charts", []) or [])
    table_count = len(getattr(sheet, "tables", {}) or {})
    if image_count:
        drawings.append(f"{image_count} images")
    if chart_count:
        drawings.append(f"{chart_count} charts")
    if table_count:
        drawings.append(f"{table_count} Excel tables")
    return drawings


def _drawing_warnings(sheet_name: str, drawings: list[str]) -> list[str]:
    return [
        f"Sheet '{sheet_name}' contains unsupported drawing content: {', '.join(drawings)}"
    ] if drawings else []


def _slug(value: str) -> str:
    normalized = "".join(char.lower() if char.isalnum() else "_" for char in value)
    normalized = "_".join(part for part in normalized.split("_") if part)
    return normalized or hashlib.sha1(value.encode("utf-8")).hexdigest()[:10]
