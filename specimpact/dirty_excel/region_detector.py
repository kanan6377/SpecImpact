from __future__ import annotations

from openpyxl.utils import get_column_letter

from specimpact.dirty_excel.cell_renderer import render_region_markdown
from specimpact.dirty_excel.models import DirtyCell, DirtyRegion, DirtySheet, RegionType


def detect_regions(sheets: list[DirtySheet], cells: list[DirtyCell]) -> list[DirtyRegion]:
    by_sheet: dict[str, list[DirtyCell]] = {sheet.sheet_id: [] for sheet in sheets}
    for cell in cells:
        if cell.value not in (None, ""):
            by_sheet.setdefault(cell.sheet_id, []).append(cell)
    regions: list[DirtyRegion] = []
    for sheet in sheets:
        sheet_cells = by_sheet.get(sheet.sheet_id, [])
        if not sheet_cells:
            continue
        for index, bounds in enumerate(_split_bounds(sheet_cells), start=1):
            start_row, end_row, start_column, end_column = bounds
            range_text = (
                f"{get_column_letter(start_column)}{start_row}:"
                f"{get_column_letter(end_column)}{end_row}"
            )
            selected = [
                cell
                for cell in cells
                if cell.sheet_id == sheet.sheet_id
                and start_row <= cell.row <= end_row
                and start_column <= cell.column <= end_column
            ]
            region = DirtyRegion(
                region_id=f"{sheet.sheet_id}.region.{index:03d}",
                workbook_id=sheet.workbook_id,
                sheet_id=sheet.sheet_id,
                sheet_name=sheet.sheet_name,
                range=range_text,
                region_type=_region_type(sheet, selected),
                rendered_text="",
                evidence_ids=[cell.evidence_id for cell in selected],
                start_row=start_row,
                end_row=end_row,
                start_column=start_column,
                end_column=end_column,
            )
            region.rendered_text = render_region_markdown(region, selected)
            regions.append(region)
    return regions


def _split_bounds(cells: list[DirtyCell]) -> list[tuple[int, int, int, int]]:
    rows = sorted({cell.row for cell in cells})
    blocks: list[list[int]] = []
    current: list[int] = []
    previous = None
    for row in rows:
        if previous is None or row <= previous + 1:
            current.append(row)
        else:
            blocks.append(current)
            current = [row]
        previous = row
    if current:
        blocks.append(current)
    bounds = []
    for block in blocks:
        block_cells = [cell for cell in cells if cell.row in block]
        columns = sorted({cell.column for cell in block_cells})
        for column_block in _split_numbers(columns):
            selected = [cell for cell in block_cells if cell.column in column_block]
            bounds.append(
                (
                    min(cell.row for cell in selected),
                    max(cell.row for cell in selected),
                    min(cell.column for cell in selected),
                    max(cell.column for cell in selected),
                )
            )
    return bounds


def _split_numbers(numbers: list[int]) -> list[list[int]]:
    blocks: list[list[int]] = []
    current: list[int] = []
    previous = None
    for number in numbers:
        if previous is None or number <= previous + 1:
            current.append(number)
        else:
            blocks.append(current)
            current = [number]
        previous = number
    if current:
        blocks.append(current)
    return blocks


def _region_type(sheet: DirtySheet, cells: list[DirtyCell]) -> RegionType:
    cell_text = " ".join(cell.value or "" for cell in cells)
    if any(token in cell_text for token in ("改訂", "履歴", "版数", "revision")):
        return "revision_history"
    validation_tokens = ("入力チェック", "チェックID", "チェック名")
    if _looks_like_screen_table(cell_text):
        return "screen_item_table"
    if any(token in cell_text for token in validation_tokens):
        return "validation_block"
    if sheet.sheet_type == "screen_item_definition" and not any(
        token in cell_text for token in validation_tokens
    ):
        return "screen_item_table"
    if sheet.sheet_type == "screen_item_definition":
        return "screen_item_table"
    if sheet.sheet_type == "validation_rule":
        return "validation_block"
    if sheet.sheet_type == "api_mapping":
        return "api_mapping_table"
    if sheet.sheet_type == "db_mapping":
        return "db_mapping_table"
    if sheet.sheet_type == "external_interface":
        return "external_if_table"
    if sheet.sheet_type == "test_case":
        return "test_case_table"
    if sheet.sheet_type == "glossary":
        return "glossary_table"
    if len(cells) <= 6:
        return "note_block"
    return "unknown"


def _looks_like_screen_table(text: str) -> bool:
    if "画面" in text and any(token in text for token in ("項目名", "物理名", "入力可否")):
        return True
    if "screen" in text.casefold() and any(token in text.casefold() for token in ("field", "item")):
        return True
    return False
