from __future__ import annotations

import html
from pathlib import Path

from openpyxl.utils import get_column_letter

from specimpact.dirty_excel.models import DirtyCell, DirtyRegion, DirtySheet, DirtyWorkbook


def render_region_markdown(region: DirtyRegion, cells: list[DirtyCell]) -> str:
    selected = [
        cell
        for cell in cells
        if cell.sheet_id == region.sheet_id
        and region.start_row <= cell.row <= region.end_row
        and region.start_column <= cell.column <= region.end_column
        and cell.value not in (None, "")
    ]
    lines = [
        f"# {region.sheet_name} {region.range}",
        "",
        "| cell | value | style_hint |",
        "| --- | --- | --- |",
    ]
    for cell in sorted(selected, key=lambda item: (item.row, item.column)):
        lines.append(
            f"| {cell.cell} | {_md(cell.value or '')} | {_md(','.join(style_hints(cell)))} |"
        )
    return "\n".join(lines)


def render_sheet_markdown(sheet: DirtySheet, cells: list[DirtyCell]) -> str:
    lines = [f"# {sheet.sheet_name}", ""]
    sheet_cells = [cell for cell in cells if cell.sheet_id == sheet.sheet_id]
    if not sheet_cells:
        return "\n".join(lines)
    max_row = max(cell.row for cell in sheet_cells)
    max_column = max(cell.column for cell in sheet_cells)
    by_pos = {(cell.row, cell.column): cell for cell in sheet_cells}
    header = ["row", *[get_column_letter(column) for column in range(1, max_column + 1)]]
    lines.append("| " + " | ".join(header) + " |")
    lines.append("| " + " | ".join("---" for _ in header) + " |")
    for row in range(1, max_row + 1):
        values = [str(row)]
        for column in range(1, max_column + 1):
            cell = by_pos.get((row, column))
            values.append(_md(cell.value if cell and cell.value is not None else ""))
        lines.append("| " + " | ".join(values) + " |")
    return "\n".join(lines)


def render_sheet_html(sheet: DirtySheet, cells: list[DirtyCell]) -> str:
    sheet_cells = [cell for cell in cells if cell.sheet_id == sheet.sheet_id]
    if not sheet_cells:
        return f"<h1>{html.escape(sheet.sheet_name)}</h1>"
    max_row = max(cell.row for cell in sheet_cells)
    max_column = max(cell.column for cell in sheet_cells)
    by_pos = {(cell.row, cell.column): cell for cell in sheet_cells}
    rows = [f"<h1>{html.escape(sheet.sheet_name)}</h1>", "<table>"]
    for row in range(1, max_row + 1):
        rows.append("<tr>")
        for column in range(1, max_column + 1):
            cell = by_pos.get((row, column))
            hints = " ".join(style_hints(cell)) if cell else ""
            value = html.escape(cell.value or "") if cell else ""
            address = html.escape(cell.cell if cell else f"{get_column_letter(column)}{row}")
            rows.append(f'<td data-cell="{address}" class="{hints}">{value}</td>')
        rows.append("</tr>")
    rows.append("</table>")
    return "\n".join(rows)


def write_rendered(
    project_root: Path,
    workbook: DirtyWorkbook,
    sheets: list[DirtySheet],
    cells: list[DirtyCell],
) -> list[Path]:
    rendered_dir = project_root / "sources" / "rendered"
    rendered_dir.mkdir(parents=True, exist_ok=True)
    paths: list[Path] = []
    for sheet in sheets:
        base = f"{workbook.workbook_id}.{sheet.sheet_index:03d}"
        markdown = rendered_dir / f"{base}.md"
        page = rendered_dir / f"{base}.html"
        markdown.write_text(render_sheet_markdown(sheet, cells), encoding="utf-8")
        page.write_text(render_sheet_html(sheet, cells), encoding="utf-8")
        paths.extend([markdown, page])
    return paths


def style_hints(cell: DirtyCell | None) -> list[str]:
    if not cell:
        return []
    hints = []
    if cell.style.font_bold:
        hints.append("bold")
    if cell.style.fill_color:
        hints.append(f"fill:{cell.style.fill_color}")
    if cell.style.border:
        hints.append("border")
    if cell.merged_range:
        hints.append(f"merged:{cell.merged_range}")
    if cell.is_hidden_row or cell.is_hidden_col:
        hints.append("hidden")
    return hints


def _md(value: str) -> str:
    return str(value).replace("|", "\\|").replace("\n", " ")
